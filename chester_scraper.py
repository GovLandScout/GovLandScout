"""
GovLandScout - Chester County (PA) Upset Tax Sale Scraper

Same low-risk shape as montco_scraper.py -- a plain chesco.org page with no
bot defense, one document, one fetch-parse-store pass. The one real
difference: Chester publishes its "Advertising List" (the legally-required
published list of upset sale properties) as an XLSX spreadsheet, not a
PDF, so this uses openpyxl instead of pdfplumber. Discovery is still
dynamic -- the document id in the URL (/DocumentCenter/View/85421 for
2026) is a new number every year, found by matching link text mentioning
"Advertising List" rather than hardcoding the id.

No street address is published here, only a legal description ("WS OF
HILLSIDE DR", "NE COR OF E BARNARD & S WORTHINGTON") -- a directional
description relative to a road, not a street number. Unlike Montgomery's
clean postal addresses, there's nothing here worth handing to a geocoder
(no house number to resolve, and the abbreviated directional format is
inconsistent enough that even extracting a clean street name would be
guesswork). These listings are stored with address=None -- they still
show in the site's table like any other listing missing that field, just
without a map pin, the same tradeoff GovEase's township-only Monroe
County listings already make.

Confirmed against the real 2026 list before writing this: 349 rows, both
the ALTID (parcel, e.g. "1-2-63") and Customer (internal account id, e.g.
"0102_00630000") columns are 100% unique and 100% populated -- ALTID is
used as account_number since it's the more standard-looking parcel format
of the two.
"""

import re
import sqlite3
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import requests

import combined_db

BASE_URL = "https://www.chesco.org"
UPSET_SALE_PAGE_URL = f"{BASE_URL}/6010/Upset-Tax-Sale-Information"
DB_PATH = "chester_properties.db"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
}

LINK_PATTERN = re.compile(
    r'<a[^>]+href="([^"]+)"[^>]*>([^<]*Advertising List[^<]*)</a>', re.IGNORECASE,
)

HEADER_FIRST_CELL = "Customer"
EXPECTED_COLUMN_COUNT = 7


def find_advertising_list_url(html: str) -> str | None:
    match = LINK_PATTERN.search(html)
    return match.group(1) if match else None


def parse_money(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value) if value > 0 else None
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    if not cleaned or not re.match(r"^\d+(\.\d+)?$", cleaned):
        return None
    return cleaned


def parse_workbook(content: bytes) -> list[dict]:
    listings = []
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    for row in sheet.iter_rows(values_only=True):
        if len(row) < EXPECTED_COLUMN_COUNT:
            continue
        customer, altid, name, name2, legal1, legal2, price = row[:EXPECTED_COLUMN_COUNT]
        if customer == HEADER_FIRST_CELL:
            continue  # header row
        if not altid:
            continue

        description = " -- ".join(p for p in (str(legal1).strip() if legal1 else None,
                                                str(legal2).strip() if legal2 else None) if p) or None

        listings.append({
            "county": "Chester",
            "account_number": str(altid).strip(),
            "minimum_bid": parse_money(price),
            "description": description,
        })
    return listings


def init_db(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS chester_properties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_number TEXT,
            minimum_bid TEXT,
            description TEXT,
            source_url TEXT,
            first_seen TEXT,
            last_seen TEXT
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_chester_account
        ON chester_properties(account_number)
    """)
    conn.commit()


def upsert_local(conn: sqlite3.Connection, listing: dict, source_url: str):
    now = datetime.now(timezone.utc).isoformat()
    existing = conn.execute(
        "SELECT id FROM chester_properties WHERE account_number = ?",
        (listing["account_number"],),
    ).fetchone()

    fields = (listing["minimum_bid"], listing["description"], source_url)
    if existing:
        conn.execute(
            """UPDATE chester_properties SET
                minimum_bid = ?, description = ?, source_url = ?, last_seen = ?
               WHERE account_number = ?""",
            fields + (now, listing["account_number"]),
        )
    else:
        conn.execute(
            """INSERT INTO chester_properties (
                minimum_bid, description, source_url, account_number, first_seen, last_seen
               ) VALUES (?, ?, ?, ?, ?, ?)""",
            fields + (listing["account_number"], now, now),
        )
    conn.commit()


def main():
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    session = requests.Session()

    print(f"Fetching {UPSET_SALE_PAGE_URL} ...")
    page_resp = session.get(UPSET_SALE_PAGE_URL, headers=HEADERS, timeout=30)
    page_resp.raise_for_status()

    doc_url = find_advertising_list_url(page_resp.text)
    if doc_url is None:
        print("Couldn't find this year's Advertising List link -- page structure may have changed.")
        return
    if not doc_url.startswith("http"):
        doc_url = f"{BASE_URL}{doc_url}"

    print(f"Fetching {doc_url} ...")
    doc_resp = session.get(doc_url, headers=HEADERS, timeout=30)
    doc_resp.raise_for_status()

    listings = parse_workbook(doc_resp.content)
    print(f"Found {len(listings)} listing(s).")

    combined_conn = combined_db.get_connection()
    for listing in listings:
        upsert_local(conn, listing, doc_url)
        combined_db.upsert_listing(
            combined_conn,
            county=listing["county"],
            account_number=listing["account_number"],
            precinct=None,
            minimum_bid=listing["minimum_bid"],
            estimated_value=None,  # Chester doesn't publish an independent value estimate
            address=None,  # no street-level address published -- see module docstring
            description=listing["description"],
            status="Active",
            source="chesco.org",
            source_url=doc_url,
            state="PA",
            commit=False,
        )
    combined_conn.commit()
    combined_conn.close()

    conn.close()
    print(f"\n{len(listings)} listing(s) stored into {DB_PATH}")


if __name__ == "__main__":
    main()
